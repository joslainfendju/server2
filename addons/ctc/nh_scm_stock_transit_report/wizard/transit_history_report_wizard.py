# -*- coding: utf-8 -*-
##############################################################################
#
#    ITS/NH-IT
#    Copyright (C) 2018-2019 NH-IT
#
##############################################################################

from io import StringIO, BytesIO
from datetime import datetime


import base64
import datetime
import time

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
from openerp.exceptions import ValidationError

#from openerp.osv import osv,orm
from odoo import api, fields, models,_
import  logging
from odoo.exceptions import UserError

#Autres bibliotheques
import datetime



class data_wizardy(models.TransientModel):
    _name = "nh_scm_stock_transit_report.data_wizard"



    @api.model
    def _get_company_id(self):
        return self.env.user.company_id



    company_id = fields.Many2one('res.company', string='Company', required="1", default=_get_company_id)
    branch_src_id = fields.Many2one('res.branch', string='Source Branch')
    branch_dest_id = fields.Many2one('res.branch', string='Destination Branch')
    direction = fields.Selection([('any', 'Any'), ('incoming', 'Incoming to transit'), ('outgoing', 'Leaving Transit')], default='any')
    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')
    filter_by = fields.Selection([('product', 'Product'), ('category', 'Product Category')], string='Filter By',
                                 default='product')
    category_id = fields.Many2one('product.category', string='Category')
    product_ids = fields.Many2many('product.product', string='Products')
    state = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File name', readonly=True)

    @api.constrains('start_date', 'end_date')
    def check_constraint(self):
        if (self.start_date >= self.end_date):
            raise UserError(_("Start date can't not be greater  than end Date"))


    def get_product(self):
        product_pool = self.env['product.product']
        if not self.filter_by:
            product_ids = product_pool.search([('type', '!=', 'service')])
            return product_ids
        elif self.filter_by == 'product' and self.product_ids:
            return self.product_ids
        elif self.filter_by == 'category' and self.category_id:
            product_ids = product_pool.search(
                [('categ_id', 'child_of', self.category_id.id), ('type', '!=', 'service')])
            return product_ids





    @api.multi
    def get_open_quantity(self, product, location):
        start_date = str(self.start_date) + ' 00:00:00'

        return product.with_context({'location': location.id, 'to_date': start_date}).qty_available

    @api.multi
    def get_transit_move_lines(self, products, branch_src=None, branch_dest=None, direction= 'any'):
        state = 'done'
        start_date = str(self.start_date) + ' 00:00:00'
        end_date = str(self.end_date) + ' 23:59:59'
        moves =[]

        base_domain = ['&', ('state', '=', state), '&', ('product_id', 'in', [p.id for p in products]),
                       '&',('date', '>=', start_date),('date', '<=', end_date)
                      ]
        if branch_src:
            move_line_ids= self.env['stock.move.line'].search(['&',('move_id.picking_id.origin','!=', False),'!',('location_id.branch_id', '=', branch_src.id)])
            line_ids = []
            for line in move_line_ids:
                pick = self.env['stock.picking'].search([('name','=', line.move_id.picking_id.origin)])
                if pick and pick.branch_id.id == branch_src.id:
                    line_ids.append(line.id)
            if line_ids :
                base_domain = ['&', '|',('location_id.branch_id', '=', branch_src.id),('id','in', line_ids)] + base_domain
            else :
                base_domain = ['&',('location_id.branch_id', '=', branch_src.id)] + base_domain

        if branch_dest:
            base_domain = ['&', ('location_dest_id.branch_id', '=', branch_dest.id)] + base_domain
        if direction == 'any':
            base_domain = ['&','|',('location_dest_id.usage', '=', 'transit'),('location_id.usage', '=', 'transit')]+base_domain
        elif direction =='incoming' :

            base_domain = ['&',('location_dest_id.usage', '=', 'transit')] + base_domain
        else:
            base_domain = ['&', ('location_id.usage', '=', 'transit')] + base_domain

        logging.info("domaine transit est "+ str(base_domain))


        return self.env['stock.move.line'].search(base_domain)





    def get_product(self):
        product_pool = self.env['product.product']
        if not self.filter_by:
            product_ids = product_pool.search([('type', '!=', 'service')])
            return product_ids
        elif self.filter_by == 'product' and self.product_ids:
            return self.product_ids
        elif self.filter_by == 'category' and self.category_id:
            product_ids = product_pool.search(
                [('categ_id', 'child_of', self.category_id.id), ('type', '!=', 'service')])
            return product_ids





    @api.multi
    def print_pdf(self):

        if (self.start_date >= self.end_date):
            raise UserError(_("Start date can't not be greater  than end Date"))

       # if not self.branch_src_id and not self.branch_dest_id:
       #     raise UserError(_("You must at least choose the destination or origin branch"))

        data = self.read()
        datas = {
            'form': self.id
        }
        return self.env.ref('nh_scm_stock_transit_report.print_stock_transit_report').report_action(self, data=datas)

    @api.multi
    def exporter_sous_excel(self):

        if (self.start_date >= self.end_date):
            raise UserError(_("Start date can't not be greater  than end Date"))

        # if not self.branch_src_id and not self.branch_dest_id:
        #    raise UserError(_("You must at least choose the destination or origin branch"))


        # initialisation du style
        couleurGrandTitre = PatternFill(start_color='2F75B5',
                                        end_color='2F75B5',
                                        fill_type='solid'
                                        )

        couleurTitre = PatternFill(start_color='FFE2EFDA',
                                   end_color='FFE2EFDA',
                                   fill_type='solid')

        couleurBranche = PatternFill(start_color='FF2F75B5',
                                     end_color='FF2F75B5',
                                     fill_type='solid')



        res = []

        # initialisation des données pour Excel
        ln = 9
        path = get_module_resource('nh_scm_stock_transit_report', 'static/template/stock_transit_report.xlsx')
        xfile = openpyxl.load_workbook(path)
        sheet = xfile.get_sheet_by_name('bilan')

        sheet['F' + str(2)] = self.env.user.name
        sheet['H' + str(2)] =str(datetime.datetime.now())

        sheet['E' + str(4)] = self.branch_src_id.name or ""
        sheet['F' + str(4)] = self.branch_dest_id.name or ""
        sheet['G' + str(4)] = self.start_date
        sheet['H' + str(4)] = self.end_date

        # debut du traitement pour chaque entrepôt de la liste des entrepôts

        feuille_courante = xfile.active

        move_lines = self.get_transit_move_lines(self.get_product(), self.branch_src_id, self.branch_dest_id, self.direction)

        for line in move_lines:

            sheet['A' + str(ln)] = line.date
            sheet['B' + str(ln)] = line.product_id.code
            sheet['C' + str(ln)] = line.product_id.product_tmpl_id.name
            sheet['D' + str(ln)] = line.lot_id.name
            sheet['E' + str(ln)] = line.product_uom_id.name

            if line.move_id and line.move_id.picking_id and line.move_id.picking_id.request_id :
                sheet['F' + str(ln)] = line.move_id.picking_id.request_id.name
            elif line.move_id and line.move_id.picking_id and line.move_id.picking_id.receipt_from_request_id:
                sheet['F' + str(ln)] = line.move_id.picking_id.receipt_from_request_id.name
            elif line.move_id and line.move_id.picking_id and line.move_id.picking_id.receipt_from_return_request_id:
                sheet['F' + str(ln)] = line.move_id.picking_id.receipt_from_return_request_id.name
            elif line.move_id and line.move_id.picking_id and line.move_id.picking_id.return_request_id:
                sheet['F' + str(ln)] = line.move_id.picking_id.return_request_id.name

            if line.move_id and line.move_id.picking_id:
                sheet['G' + str(ln)] = line.move_id.picking_id.name

            sheet['H' + str(ln)] = line.location_id.display_name
            sheet['I' + str(ln)] = line.location_dest_id.display_name

            transit_location = None

            if line.location_dest_id.usage == 'transit':
                sheet['J' + str(ln)] = 'Incoming to transit'
                transit_location = line.location_dest_id
            else:
                sheet['j' + str(ln)] = 'Leaving to transit'
                transit_location = line.location_id


            sheet['K' + str(ln)] = line.qty_done

            sheet['L' + str(ln)] = line.product_id.uom_po_id._compute_quantity(line.qty_done, line.product_uom_id,
                                                                                rounding_method='HALF-UP') * line.product_id.standard_price
            ln =ln +1




        current_date = time.strftime("%Y_%m_%d")

        # nomFichier='RAPPORT_credit_limit'+'_'+str(current_date)+'.xlsx'
        name = u'STOCK transit REPORT-' + u'V-' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ".xlsx"
        # output = StringIO()
        # xfile.save(output)

        current_date = time.strftime("%Y_%m_%d")
        chemin = get_module_resource('nh_scm_stock_transit_report', 'static', 'template')
        nomFichier = chemin + '_' + str(current_date) + '.xlsx'
        xfile.save(nomFichier)
        self.name = name
        encodedFile = None
        with open(nomFichier, "rb") as f:
            encodedFile = base64.b64encode(f.read())

        data = encodedFile
        self.write({

                    'state': 'get',
                    'data': data,
                    'name': name
                   })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nh_scm_stock_transit_report.data_wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',

        }


class stock_transit_report_excel(models.TransientModel):
    _name = "nh_scm.stock_transit_report_excel"

    excel_file = fields.Binary('Excel Report')
    file_name = fields.Char('Excel File')

