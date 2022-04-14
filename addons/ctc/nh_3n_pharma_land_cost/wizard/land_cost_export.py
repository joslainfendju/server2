# -*- encoding: utf-8 -*-

import base64
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
from openerp.exceptions import ValidationError

#from openerp.osv import osv,orm
from odoo import api, fields, models,_
import  logging
from odoo.exceptions import UserError

#Autres bibliotheques
import datetime
_logger = logging.getLogger(__name__)


class LandCostExport(models.TransientModel):
    _name = 'nh_3n_pharma_land_cost.land_cost_export_wizard'

    #
    land_cost_id = fields.Many2one('nh_3n_pharma_land_cost.land_cost', 'Landing Cost')

    state = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File name', readonly=True)


    @api.multi
    def button_print_excel(self):
        # initialisation du style

        couleurFormula = PatternFill(start_color='C4D79B',
                                     end_color='C4D79B',
                                     fill_type='solid')

        couleurManual = PatternFill(start_color='FFFF00',
                                    end_color='FFFF00',
                                    fill_type='solid')

        couleurDifference = PatternFill(start_color='FFFFFF',
                                    end_color='FFFFFF',
                                    fill_type='solid')

        # initialisation des données pour Excel

        path = get_module_resource('nh_3n_pharma_land_cost', 'static/template/land_cost_template.xlsx')
        xfile = openpyxl.load_workbook(path)
        sheet_calcul = xfile.get_sheet_by_name('CALCUL')
        sheet_resume = xfile.get_sheet_by_name('RESUME')
        sheet_final = xfile.get_sheet_by_name('FINAL')

        # remplissage des entêtes

        sheet_calcul['B4'] = self.land_cost_id.order_id.company_id.name
        sheet_calcul['E4'] = self.land_cost_id.order_id.branch_id.name
        sheet_calcul['I4'] = self.land_cost_id.order_id.name
        sheet_calcul['L4'] = self.land_cost_id.name
        sheet_calcul['N4'] = self.land_cost_id.create_date

        # debut du traitement pour chaque ligne de calcul de prix

        ln = 10

        ln_final = 6
        total_qty = 0
        total_unit_price = 0
        total_ratio = 0
        total_ratio_volume = 0
        for line in self.land_cost_id.land_cost_line_ids:
            sheet_calcul['A' + str(ln)] = line.product_id.code
            sheet_calcul['B' + str(ln)] = line.product_id.product_tmpl_id.name
            sheet_calcul['C' + str(ln)] = line.product_id.uom_id.name
            sheet_calcul['D' + str(ln)] = line.purchase_qty
            sheet_calcul['E' + str(ln)] = line.purchase_unit_price
            sheet_calcul['F' + str(ln)] = line.ratio
            sheet_calcul['H' + str(ln)] = line.purchase_price
            sheet_calcul['I' + str(ln)] = line.line_custom_cost
            sheet_calcul['J' + str(ln)] = line.line_outlay_charge
            sheet_calcul['K' + str(ln)] = line.line_intervention_charge
            sheet_calcul['L' + str(ln)] = line.line_freight
            sheet_calcul['M' + str(ln)] = line.line_financial_expense_value
            sheet_calcul['N' + str(ln)] = line.line_whithholding_on_sale_value
            sheet_calcul['O' + str(ln)] = line.line_handling_cost
            sheet_calcul['P' + str(ln)] = line.line_external_audit_cost_value
            sheet_calcul['Q' + str(ln)] = line.global_cost_purchase
            sheet_calcul['R' + str(ln)] = line.cost_per_unit_purchase
            sheet_calcul['S' + str(ln)] = line.stock_before_arrival
            sheet_calcul['T' + str(ln)] = line.cost_before_purchase
            sheet_calcul['U' + str(ln)] = line.cost_price_after_purchase

            sheet_final['A' + str(ln_final)] = line.product_id.code
            sheet_final['B' + str(ln_final)] = line.product_id.product_tmpl_id.name
            sheet_final['C' + str(ln_final)] = line.product_id.uom_id.name
            sheet_final['D' + str(ln_final)] = line.cost_price_after_purchase

            if line.use_ratio_volume:
                sheet_calcul['G' + str(ln)] = line.volume_ratio
                total_ratio_volume += line.volume_ratio
            else:
                sheet_calcul['G' + str(ln)] = line.ratio
                total_ratio_volume += line.ratio

            total_qty += line.purchase_qty
            total_unit_price += line.purchase_unit_price
            total_ratio += line.ratio

            ln += 1
            ln_final += 1

        sheet_calcul['A' + str(ln)].fill = couleurFormula
        sheet_calcul['A' + str(ln)] = "Formula Total"
        sheet_calcul['B' + str(ln)].fill = couleurFormula
        # sheet_calcul['B' + str(ln)] = line.product_id.product_tmpl_id.name
        sheet_calcul['C' + str(ln)].fill = couleurFormula
        # sheet_calcul['C' + str(ln)] = line.product_id.uom_id.name
        sheet_calcul['D' + str(ln)].fill = couleurFormula
        sheet_calcul['D' + str(ln)] = total_qty
        sheet_calcul['E' + str(ln)].fill = couleurFormula
        sheet_calcul['E' + str(ln)] = total_unit_price
        sheet_calcul['F' + str(ln)].fill = couleurFormula
        sheet_calcul['F' + str(ln)] = total_ratio
        sheet_calcul['G' + str(ln)].fill = couleurFormula
        sheet_calcul['G' + str(ln)] = total_ratio_volume
        sheet_calcul['H' + str(ln)].fill = couleurFormula
        sheet_calcul['H' + str(ln)] = self.land_cost_id.amount_total_formula
        sheet_calcul['I' + str(ln)].fill = couleurFormula
        sheet_calcul['I' + str(ln)] = self.land_cost_id.custom_cost_formula
        sheet_calcul['J' + str(ln)].fill = couleurFormula
        sheet_calcul['J' + str(ln)] = self.land_cost_id.outlay_charge_formula
        sheet_calcul['K' + str(ln)].fill = couleurFormula
        sheet_calcul['K' + str(ln)] = self.land_cost_id.intervention_charge_formula
        sheet_calcul['L' + str(ln)].fill = couleurFormula
        sheet_calcul['L' + str(ln)] = self.land_cost_id.freight_formula
        sheet_calcul['M' + str(ln)].fill = couleurFormula
        sheet_calcul['M' + str(ln)] = self.land_cost_id.financial_expense_value_formula
        sheet_calcul['N' + str(ln)].fill = couleurFormula
        sheet_calcul['N' + str(ln)] = self.land_cost_id.whithholding_on_sale_value_formula
        sheet_calcul['O' + str(ln)].fill = couleurFormula
        sheet_calcul['O' + str(ln)] = self.land_cost_id.handling_cost_formula
        sheet_calcul['P' + str(ln)].fill = couleurFormula
        sheet_calcul['P' + str(ln)] = self.land_cost_id.external_audit_cost_value_formula
        sheet_calcul['Q' + str(ln)].fill = couleurFormula
        sheet_calcul['Q' + str(ln)] = self.land_cost_id.amount_total_with_land_costs_formula
        sheet_calcul['R' + str(ln)].fill = couleurFormula
        sheet_calcul['S' + str(ln)].fill = couleurFormula
        sheet_calcul['T' + str(ln)].fill = couleurFormula
        sheet_calcul['U' + str(ln)].fill = couleurFormula

        ln+=1

        sheet_calcul['A' + str(ln)].fill = couleurManual
        sheet_calcul['A' + str(ln)] = "Manual Total"
        sheet_calcul['B' + str(ln)].fill = couleurManual
        # sheet_calcul['B' + str(ln)] = line.product_id.product_tmpl_id.name
        sheet_calcul['C' + str(ln)].fill = couleurManual
        # sheet_calcul['C' + str(ln)] = line.product_id.uom_id.name
        sheet_calcul['D' + str(ln)].fill = couleurManual
        sheet_calcul['D' + str(ln)] = total_qty
        sheet_calcul['E' + str(ln)].fill = couleurManual
        sheet_calcul['E' + str(ln)] = total_unit_price
        sheet_calcul['F' + str(ln)].fill = couleurManual
        sheet_calcul['F' + str(ln)] = total_ratio
        sheet_calcul['G' + str(ln)].fill = couleurManual
        sheet_calcul['G' + str(ln)] = total_ratio_volume
        sheet_calcul['H' + str(ln)].fill = couleurManual
        sheet_calcul['H' + str(ln)] = self.land_cost_id.amount_total
        sheet_calcul['I' + str(ln)].fill = couleurManual
        sheet_calcul['I' + str(ln)] = self.land_cost_id.custom_cost
        sheet_calcul['J' + str(ln)].fill = couleurManual
        sheet_calcul['J' + str(ln)] = self.land_cost_id.outlay_charge
        sheet_calcul['K' + str(ln)].fill = couleurManual
        sheet_calcul['K' + str(ln)] = self.land_cost_id.intervention_charge
        sheet_calcul['L' + str(ln)].fill = couleurManual
        sheet_calcul['L' + str(ln)] = self.land_cost_id.freight
        sheet_calcul['M' + str(ln)].fill = couleurManual
        sheet_calcul['M' + str(ln)] = self.land_cost_id.financial_expense_value
        sheet_calcul['N' + str(ln)].fill = couleurManual
        sheet_calcul['N' + str(ln)] = self.land_cost_id.whithholding_on_sale_value
        sheet_calcul['O' + str(ln)].fill = couleurManual
        sheet_calcul['O' + str(ln)] = self.land_cost_id.handling_cost
        sheet_calcul['P' + str(ln)].fill = couleurManual
        sheet_calcul['P' + str(ln)] = self.land_cost_id.external_audit_cost_value
        sheet_calcul['Q' + str(ln)].fill = couleurManual
        sheet_calcul['Q' + str(ln)] = self.land_cost_id.amount_total_with_land_costs
        sheet_calcul['R' + str(ln)].fill = couleurManual
        sheet_calcul['S' + str(ln)].fill = couleurManual
        sheet_calcul['T' + str(ln)].fill = couleurManual
        sheet_calcul['U' + str(ln)].fill = couleurManual
        
        
        ln += 1
        

        sheet_calcul['A' + str(ln)].fill = couleurDifference
        sheet_calcul['A' + str(ln)] = "Difference Total"
        sheet_calcul['B' + str(ln)].fill = couleurDifference
        # sheet_calcul['B' + str(ln)] = line.product_id.product_tmpl_id.name
        sheet_calcul['C' + str(ln)].fill = couleurDifference
        # sheet_calcul['C' + str(ln)] = line.product_id.uom_id.name
        sheet_calcul['D' + str(ln)].fill = couleurDifference
        #sheet_calcul['D' + str(ln)] = total_qty
        sheet_calcul['E' + str(ln)].fill = couleurDifference
        #sheet_calcul['E' + str(ln)] = total_unit_price
        sheet_calcul['F' + str(ln)].fill = couleurDifference
        sheet_calcul['F' + str(ln)] = 100 - total_ratio
        sheet_calcul['G' + str(ln)].fill = couleurDifference
        sheet_calcul['G' + str(ln)] = 100 - total_ratio_volume
        sheet_calcul['H' + str(ln)].fill = couleurDifference
        sheet_calcul['H' + str(ln)] = self.land_cost_id.amount_total_difference
        sheet_calcul['I' + str(ln)].fill = couleurDifference
        sheet_calcul['I' + str(ln)] = self.land_cost_id.custom_cost_difference
        sheet_calcul['J' + str(ln)].fill = couleurDifference
        sheet_calcul['J' + str(ln)] = self.land_cost_id.outlay_charge_difference
        sheet_calcul['K' + str(ln)].fill = couleurDifference
        sheet_calcul['K' + str(ln)] = self.land_cost_id.intervention_charge_difference
        sheet_calcul['L' + str(ln)].fill = couleurDifference
        sheet_calcul['L' + str(ln)] = self.land_cost_id.freight_difference
        sheet_calcul['M' + str(ln)].fill = couleurDifference
        sheet_calcul['M' + str(ln)] = self.land_cost_id.financial_expense_value_difference
        sheet_calcul['N' + str(ln)].fill = couleurDifference
        sheet_calcul['N' + str(ln)] = self.land_cost_id.whithholding_on_sale_value_difference
        sheet_calcul['O' + str(ln)].fill = couleurDifference
        sheet_calcul['O' + str(ln)] = self.land_cost_id.handling_cost_difference
        sheet_calcul['P' + str(ln)].fill = couleurDifference
        sheet_calcul['P' + str(ln)] = self.land_cost_id.external_audit_cost_value_difference
        sheet_calcul['Q' + str(ln)].fill = couleurDifference
        sheet_calcul['Q' + str(ln)] = self.land_cost_id.amount_total_with_land_costs_difference
        sheet_calcul['R' + str(ln)].fill = couleurDifference
        sheet_calcul['S' + str(ln)].fill = couleurDifference
        sheet_calcul['T' + str(ln)].fill = couleurDifference
        sheet_calcul['U' + str(ln)].fill = couleurDifference


        sheet_resume['d3'] = self.land_cost_id.custom_cost
        sheet_resume['d4'] = self.land_cost_id.intervention_charge
        sheet_resume['d5'] = self.land_cost_id.outlay_charge
        sheet_resume['d6'] = self.land_cost_id.freight


        current_date = time.strftime("%Y_%m_%d")

        # nomFichier='RAPPORT_credit_limit'+'_'+str(current_date)+'.xlsx'
        self.name = u'LANDING COST TEMPLATE-' + u'V-' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ".xlsx"
        # output = StringIO()
        # xfile.save(output)

        current_date = time.strftime("%Y_%m_%d")
        chemin = get_module_resource('nh_3n_pharma_land_cost', 'static', 'output')
        nomFichier = chemin + '_' + str(current_date) + '.xlsx'
        xfile.save(nomFichier)

        encodedFile = None
        with open(nomFichier, "rb") as f:
            encodedFile = base64.b64encode(f.read())

        data = encodedFile
        self.write({'state': 'get',
                    'data': data,
                    'name': self.name,
                    })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nh_3n_pharma_land_cost.land_cost_export_wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',

        }