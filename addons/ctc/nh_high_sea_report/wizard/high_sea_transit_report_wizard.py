# -*- coding: utf-8 -*-
##############################################################################
#
#    ITS/NH-IT
#    Copyright (C) 2018-2019 NH-IT
#
##############################################################################

from io import StringIO, BytesIO
from datetime import datetime


import base64
import datetime
import time

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
from openerp.exceptions import ValidationError

#from openerp.osv import osv,orm
from odoo import api, fields, models,_
import  logging
from odoo.exceptions import UserError

#Autres bibliotheques
import datetime



class data_wizardy(models.TransientModel):
    _name = "nh_high_sea_report.rt_wizard"



    @api.model
    def _get_company_id(self):
        return self.env.user.company_id


    company_id = fields.Many2one('res.company', string='Company', required="1", default=_get_company_id)
    country_src_id = fields.Many2one('res.country', string='Country of Embarkation')
    country_dest_id = fields.Many2one('res.branch', string='Landing country')
    when = fields.Selection([('current', 'At the moment'), ('date', 'On a specific Date')], default='current')
    specific_date = fields.Date('As on date')

    filter_by = fields.Selection([('product', 'Product'), ('category', 'Product Category')], string='Filter By',
                                 default='product')
    category_id = fields.Many2one('product.category', string='Category')
    product_ids = fields.Many2many('product.product', string='Products')
    state = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File name', readonly=True)

    @api.constrains('specific_date', 'when')
    def check_constraint(self):
        if (self.when == 'date' and  not self.specific_date):
            raise UserError(_("You most enter a date when you don't choose the current moment"))


    def get_product(self):
        product_pool = self.env['product.product']
        if not self.filter_by:
            product_ids = product_pool.search([('type', '!=', 'service')])
            return product_ids
        elif self.filter_by == 'product' and self.product_ids:
            return self.product_ids
        elif self.filter_by == 'category' and self.category_id:
            product_ids = product_pool.search(
                [('categ_id', 'child_of', self.category_id.id), ('type', '!=', 'service')])
            return product_ids







    @api.multi
    def get_transit_purchase_lines(self, products, country_src=None, country_dest=None):
        date_format = "%Y-%m-%d %H:%M:%S"

        specific_date = datetime.datetime.strptime(fields.Datetime.now(), date_format).replace(hour=0, minute=0, second=0,
                                                                                       microsecond=0)
        specific_date = specific_date.strftime('%Y-%m-%d') + ' 23:59:59'
        base_domain = []
        en_transit_domain = ['&',  '!', ('order_id.state', 'in', ('done', 'cancel','landed','cost_saved',)),
                             '&',  ('product_id', 'in', [p.id for p in products]),
                                   ('order_id.date_order', '<=', specific_date)
                            ]
        if self.when == 'current':
            base_domain = en_transit_domain

        else:
            specific_date = str(self.specific_date) + ' 23:59:59'
            base_domain = ['&', ('order_id.date_order', '<=', specific_date),
                           '&', ('product_id', 'in', [p.id for p in products]),

                              '|','&',('order_id.state', 'in', ('done','landed','cost_saved',)), ('order_id.date_arrival','>=',specific_date),
                                  '&',('order_id.state', 'in', ('cancel',)), ('order_id.write_date','>=',specific_date)

                           ]
            en_transit_domain = [
                                    '&', ('order_id.state', 'in', ('done', 'cancel','landed','cost_saved',)),
                                    '&', ('product_id', 'in', [p.id for p in products]),('order_id.date_order', '<=', specific_date)
                                 ]
            base_domain = ['|',] + base_domain+en_transit_domain


        if country_src:
            base_domain = ['&',
                           ('order_id.origin_country.id', '=' , country_src.id)
                          ] + base_domain

        if country_dest:
            base_domain = ['&',
                           ('order_id.origin_country.id', '=', country_dest.id)
                           ] + base_domain

        base_domain = ['&', ('order_id.type','=','abroad') ]+base_domain
        logging.info("domaine transit dans stock real transit report est "+ str(base_domain))



        return self.env['purchase.order.line'].search(base_domain, order='order_id ASC')




    def get_product(self):
        product_pool = self.env['product.product']
        if not self.filter_by:
            product_ids = product_pool.search([('type', '!=', 'service')])
            return product_ids
        elif self.filter_by == 'product' and self.product_ids:
            return self.product_ids
        elif self.filter_by == 'category' and self.category_id:
            product_ids = product_pool.search(
                [('categ_id', 'child_of', self.category_id.id), ('type', '!=', 'service')])
            return product_ids





    @api.multi
    def print_pdf(self):

        date_format = "%Y-%m-%d %H:%M:%S"
        specific_date = fields.Datetime.now()

        if self.specific_date:
            specific_date = self.specific_date

        datas = {

            'country_src_id':self.country_src_id.name,
            'country_dest_id': self.country_dest_id.name,
            'specific_date' : specific_date

        }

        elements =  self.get_json_lines()



        datas.update({

            'ids': self.ids,
            'elements': elements,
            'model': 'nh_high_sea_report.report_transit'

        })

        return {
            'type': 'ir.actions.report',
            'name': 'nh_high_sea_report.report_transit',
            'res_model': 'nh_high_sea_report.report_transit',
            'model': 'report.nh_high_sea_report.report_transit',
            'report_type': 'qweb-pdf',
            'report_name': 'nh_high_sea_report.report_transit',
            'data': datas,

        }

    @api.multi
    def exporter_sous_excel(self):


        if (self.when == 'date' and  not self.specific_date):
            raise UserError(_("You most enter a date when you don't choose the current moment"))

        elements = self.get_json_lines()




        # initialisation du style
        couleurGrandTitre = PatternFill(start_color='2F75B5',
                                        end_color='2F75B5',
                                        fill_type='solid'
                                        )

        couleurTitre = PatternFill(start_color='FFE2EFDA',
                                   end_color='FFE2EFDA',
                                   fill_type='solid')

        couleurBranche = PatternFill(start_color='FF2F75B5',
                                     end_color='FF2F75B5',
                                     fill_type='solid')



        res = []

        # initialisation des données pour Excel
        ln = 8
        path = get_module_resource('nh_high_sea_report', 'static/template/high_sea_transit_report.xlsx')
        logging.info("chemin :"+ str(path))
        xfile = openpyxl.load_workbook(path)
        sheet = xfile.get_sheet_by_name('bilan')



        sheet['D' + str(3)] = self.country_src_id.name or ""
        sheet['E' + str(3)] = self.country_dest_id.name or ""

        if self.specific_date:
            sheet['F' + str(3)] = self.specific_date
        else:
            sheet['F' + str(3)] = fields.Datetime.now()

        sheet['G' + str(3)] = self.env.user.name
        sheet['H' + str(3)] = str(datetime.datetime.now())

        # debut du traitement pour chaque entrepôt de la liste des entrepôts

        feuille_courante = xfile.active



        for line in elements:

            sheet['A' + str(ln)] = line['order']
            sheet['B' + str(ln)] = line['date']
            sheet['C' + str(ln)] = line['value']
            sheet['D' + str(ln)] = line['bill']
            sheet['E' + str(ln)] = line['state']
            sheet['F' + str(ln)] = line['sgs']
            sheet['G' + str(ln)] = line['origin_country']
            sheet['H' + str(ln)] = line['arrival_country']
            sheet['I' + str(ln)] = line['eta']
            sheet['J' + str(ln)] = line['delays']
            ln =ln +1




        current_date = time.strftime("%Y_%m_%d")

        # nomFichier='RAPPORT_credit_limit'+'_'+str(current_date)+'.xlsx'
        name = u'HIGH SEA Transit REPORT-' + u'V-' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ".xlsx"
        # output = StringIO()
        # xfile.save(output)

        current_date = time.strftime("%Y_%m_%d")
        chemin = get_module_resource('nh_high_sea_report', 'static', 'template')
        nomFichier = chemin + '_' + str(current_date) + '.xlsx'
        xfile.save(nomFichier)
        self.name = name
        encodedFile = None
        with open(nomFichier, "rb") as f:
            encodedFile = base64.b64encode(f.read())

        data = encodedFile
        self.write({

                    'state': 'get',
                    'data': data,
                    'name': name
                   })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nh_high_sea_report.rt_wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',

        }

    @api.multi
    def get_json_lines(self):

        res = []
        lines = self.get_transit_purchase_lines(self.get_product(), self.country_src_id, self.country_dest_id)
        date_format = "%Y-%m-%d %H:%M:%S"

        consume_ids= []
        i = 0
        n = len(lines)
        while i < n :

            line = lines [i]
            vals = {}
            if line.order_id.expected_landed_date :
                eta = datetime.datetime.strptime(line.order_id.expected_landed_date+' 00:00:00', date_format)
            else :
                logging.info('eta : '+line.order_id.date_order)
                eta = datetime.datetime.strptime(line.order_id.date_order, date_format)

            specific_date = datetime.datetime.strptime(fields.Datetime.now(), date_format).replace(hour=0, minute=0,
                                                                                               second=0,
                                                                                               microsecond=0)
            if self.when != 'current':
                specific_date = datetime.datetime.strptime(self.specific_date + ' 00:00:00', date_format).replace(
                    hour=0, minute=0,
                    second=0,
                    microsecond=0)


            dateDone = None
            dateDoneString = None

            if line.order_id.state in ('cancel' , 'done',) :
                dateDoneString = line.order_id.write_date
            elif line.order_id.state == 'landed' :
                if line.order_id.date_arrival:
                    dateDoneString = line.order_id.date_arrival+' 00:00:00'
                else:
                    dateDoneString = line.order_id.write_date
            elif line.order_id.state == 'cost_saved' :
                if line.order_id.date_cost_saved:
                    dateDoneString = line.order_id.date_cost_saved+' 00:00:00'
                else:
                    dateDoneString = line.order_id.write_date


            if dateDoneString :
                dateDone = datetime.datetime.strptime(dateDoneString, date_format).replace(
                    hour=0, minute=0,
                    second=0,
                    microsecond=0)

            eta = eta.replace(
                hour=0, minute=0,
                second=0,
                microsecond=0)



            delays = 0
            if dateDone:
                if dateDone > specific_date :
                    dateDone = specific_date
                delays = max ((dateDone - eta).total_seconds() / 86400,0)
            else:
                delays = max((specific_date - eta).total_seconds() / 86400, 0)



            vals.update({'order': line.order_id.name})
            vals.update({'date':  line.order_id.date_order})
            vals.update({'bill':  line.order_id.bill_of_landing_number or ''})
            vals.update({'state':  line.order_id.state})
            vals.update({'origin_country':  line.order_id. origin_country and line.order_id. origin_country.name or ''})
            vals.update({'arrival_country':  line.order_id.arrival_country and line.order_id.arrival_country.name or ''})
            vals.update({'sgs':  line.order_id.sgs_number or '' })
            vals.update({'eta':  line.order_id.expected_landed_date })
            vals.update({'value': line.order_id.amount_total})
            vals.update({'delays': str(delays)})
            res.append(vals)
            i=i+1
            while i < n and lines[i].order_id.id == line.order_id.id:
                i = i + 1

        return res


class real_time_stock_transit_report_excel(models.TransientModel):
    _name = "nh_scm.real_time_stock_transit_report_excel"

    excel_file = fields.Binary('Excel Report')
    file_name = fields.Char('Excel File')

