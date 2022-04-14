# -*- coding: utf-8 -*-

#bibliotheques pour EXCEL
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
#fin bibliotheques pour EXCEL

#Bibliotheques pour le modèle
from odoo import models, fields, api
#pour les logs
import logging
_logger = logging.getLogger(__name__)


#pour les fichiers
import base64
#Autres bibliotheques
import datetime
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, make_msgid, formatdate, getaddresses, formataddr

import os
import openpyxl
from os import listdir
import openerp.exceptions
from os.path import isfile, join





class Contact(models.Model):
    _name = 'nh_alert_contact.contact'
    _inherit = 'nh_alert_contact.contact'

    alert_transit_stock = fields.Boolean(string='Alert-transit-stock',default=False)

    


class StockQuant(models.Model):
    _inherit=['stock.quant']


   
    def analyserEtAlerterStockTrannsit(self):
        #logging.info("##debut de l'analyse des  futurs perimes ou permises###")
        #logging.info("...Chargement des parametres systemes...")

        branches = self.env['res.branch'].search([('id','>',0)])

        for branch in branches:

            duree_critique_en_transit = self.sudo().env['ir.config_parameter'].search([('key', '=', 'DUREE_CRITIQUE_STOCK_EN_TRANSIT_EN_JOUR')])

            duree_critique_en_transit_valeur = float(duree_critique_en_transit.value)

            quants = self.env['stock.quant'].search(['&',('location_id.branch_id', '=', branch.id) , ('location_id.usage', '=','transit')])


            date_format = "%Y-%m-%d %H:%M:%S"
            seuil_alerte=duree_critique_en_transit_valeur*24*3600#on convertit la valeur du paramètre en secondes

            #initialisation des données pour Excel
            ln = 7
            path = get_module_resource('nh_transit_periodic_alert', 'static/template/transit_template.xlsx')
            xfile = openpyxl.load_workbook(path)
            sheet = xfile.get_sheet_by_name('rapport')
            feuille_courante=xfile.active

            #Initialisation du style
            rouge = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')

            orange = PatternFill(start_color='FFFF8000',
                       end_color='FFFF8000',
                       fill_type='solid')
            sheet['G2'] = branch.name

            sheet['G3'] = str(datetime.datetime.now())
            sheet['G4'] = duree_critique_en_transit_valeur
            #Fin Initialisation du style





            #logging.info("...DEBUT DE L'ANALYSE...")
            maintenant=datetime.datetime.today().replace(hour=0,minute=0,second=0,microsecond=0)

            for quant in quants:
                serie = quant.lot_id
                #logging.info("Analyse d'un numero de serie %s",serie.display_name)
                alerter=None

                dateEntree = datetime.datetime.strptime(quant.in_date, date_format).replace(hour=0,minute=0,second=0,microsecond=0)
                diff = (maintenant-dateEntree).total_seconds()

                if(diff < seuil_alerte):
                    alerter=False

                else:
                    alerter=True
                    sheet['B' + str(ln)] = serie.product_id.code
                    sheet['C' + str(ln)] = serie.product_id.product_tmpl_id.name
                    sheet['D' + str(ln)] = serie.name
                    sheet['E' + str(ln)] = serie.product_id.standard_price
                    sheet['F' + str(ln)] = quant.location_id.complete_name
                    sheet['G' + str(ln)] = quant.quantity
                    sheet['H' + str(ln)] = quant.product_id.uom_id.name
                    sheet['I' + str(ln)] = serie.product_id.product_tmpl_id.standard_price * quant.quantity
                    tps = diff/86400 #conversion des sécondes en jours
                    sheet['j' + str(ln)] = quant.in_date
                    sheet['k' + str(ln)] = tps
                    ln = ln+1


            #Sauvegarde des données dans le repertoire approprié

            current_date=time.strftime("%Y_%m_%d")
            chemin=get_module_resource('nh_transit_periodic_alert', 'static', 'product_delay_in_transit_report')
            nomFichier=chemin+'_'+branch.name+'_'+str(current_date)+'.xlsx'
            xfile.save(nomFichier)


            #préparation et envoie des mails

            subject="Liste des produits qui ont mis longtemps en transit "+time.strftime("%d/%m/%Y %H:%M:%S")
            body_text=""" 
                        Bonjour A vous.
                        Vous trouverez ci-joint, le rapport des articles qui ont duré longtemps en transit.
                        Ces produits ont faits plus de %s jours en transit.
                        Cordialement.
                      """%(duree_critique_en_transit_valeur)

            destinataires = self.env['nh_alert_contact.contact'].search([ '&',('alert_transit_stock','=',True) ,('active','=',True)])
            destinataires = destinataires.filtered(lambda x: branch.id in [b.id for b in x.branch_ids])
            recipients=[]
            for d in destinataires:
                recipients.append(d.email)
            if recipients:
                sender="3n_pharma"
                outer = MIMEMultipart()
                outer['Subject'] = subject
                outer['To'] = str(COMMASPACE.join(recipients))
                outer['From'] = sender
                outer['Date'] = formatdate(localtime=True)
                email_body = MIMEText(body_text)
                outer.attach(email_body)
                files=[nomFichier]
                for f in files:
                    msg=MIMEBase('application', "octet-stream")
                    msg.set_payload(open(f,"rb").read())
                    encoders.encode_base64(msg)
                    msg.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(f))
                    outer.attach(msg)

                composed=outer.as_string()



                mail_serverObj=self.env['ir.mail_server']
                for objMail in mail_serverObj.search([('active','=',True)]):

                    smtp=False
                    try:
                        smtp = mail_serverObj.connect(objMail.smtp_host, objMail.smtp_port, objMail.smtp_user, objMail.smtp_pass, objMail.smtp_encryption or False, objMail.smtp_debug)
                        smtp.sendmail(sender,recipients, composed)

                        smtp.quit()
                    except Exception as e:
                        logging.info("Echec  :%s",e)


        

