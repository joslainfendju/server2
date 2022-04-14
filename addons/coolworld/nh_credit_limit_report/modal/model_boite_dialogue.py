# -*- encoding: utf-8 -*-

import base64
import datetime
import time
# from xlwt import Workbook
# import xlwt
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
from openerp.exceptions import ValidationError

#from openerp.osv import osv,orm
from odoo import api, fields, models,_
import  logging
from odoo.exceptions import UserError

#Autres bibliotheques
import datetime



#_logger = logging.getLogger(__name__)

class CreditLimitReport(models.TransientModel):
    _name= 'nh_credit_limit_report.credit_limit_report_wizard'


   
    # l'entité pour laquelle le rapport doit être fait, sinon précisé, le traitement s'applique à tous les entrepôts
    branch_id=fields.Many2one('res.branch','branch')
    

    state=fields.Selection([('choose','choose'),('get','get')],default='choose')
    data=fields.Binary('File',readonly=True)
    name=fields.Char('File name',readonly=True)
   
    

    def generer_fichier(self):
        #initialisation des repositories de l'ORM
        branch_obj = self.env['res.branch']
        customer_obj = self.env['res.partner']
        

        #initialisation du style
        couleurTitre = PatternFill(start_color='FFE2EFDA',
                   end_color='FFE2EFDA',
                   fill_type='solid')


        couleurBranche = PatternFill(start_color='FF2F75B5',
                   end_color='FF2F75B5',
                   fill_type='solid')


        
        #recherche des entrepôts concernés
        if(self.branch_id):
            branches=[self.branch_id]
          
        else:
            branches=branch_obj.search([('id','>',0)])

            



        #initialisation des données pour Excel
        ln =8
        path = get_module_resource('nh_credit_limit_report', 'static/template/customer_credit_limit.xlsx')
        xfile = openpyxl.load_workbook(path)
        sheet = xfile.get_sheet_by_name('bilan')

        #remplissage des entêtes
        if(self.branch_id):
            sheet['E' + str(3)] = self.branch_id.name
        else:
            sheet['E' + str(3)] = 'Toutes (All)'
        
                             
        #debut du traitement pour chaque entrepôt de la liste des entrepôts
       
        feuille_courante=xfile.active 
        for i in range(len(branches)+1):
            branche=None
            if(i<len(branches)):
                branche=branches[i]
            if(not branche):
                sheet['B' + str(ln)] ="Branche inconnue (Branch Not Defined)"
            else:
                sheet['B' + str(ln)] =branche.name
            
            feuille_courante['B' + str(ln)].fill = couleurBranche
            ln=ln+1
            sheet['B' + str(ln)] = 'Code  (Code)'
            feuille_courante['B' + str(ln)].fill = couleurTitre
            

            sheet['C' + str(ln)] = 'Nom (Name)'
            feuille_courante['C' + str(ln)].fill = couleurTitre
            

            sheet['D' + str(ln)] = 'Téléphone (Phone)'
            feuille_courante['D' + str(ln)].fill = couleurTitre
            

            sheet['E' + str(ln)] = 'Adresse Mail (email)'
            feuille_courante['E' + str(ln)].fill = couleurTitre
            

            sheet['F' + str(ln)] = 'Limite de crédit (Credit Limit)'
            feuille_courante['F' + str(ln)].fill = couleurTitre

            sheet['G' + str(ln)] = 'Condition de paiement (Payment Term)'
            feuille_courante['G' + str(ln)].fill = couleurTitre

            sheet['H' + str(ln)] = 'Dettes Accumulées (Accumulated Debt)'
            feuille_courante['H' + str(ln)].fill = couleurTitre
            

            sheet['I' + str(ln)] = 'Credit restant (Remaining Credit)'
            feuille_courante['I' + str(ln)].fill = couleurTitre

           
           
            ln=ln+1
        
            
            total_branche_limite=0
            total_branche_dette=0
            total_branche_credit_disponible=0
            if(branche):
                clients=customer_obj.search(['&',('customer','=',True),('branch_id','=',branche.id)])
            else:
                clients=customer_obj.search(['&',('customer','=',True),'|',('branch_id','=',False),'|',('branch_id','=',0),('branch_id','=',None)])
           
            
            for client in clients:
                
                sheet['B' + str(ln)] = client.id or ''
                sheet['C' + str(ln)] = client.name or ''
                sheet['D' + str(ln)] = client.phone or ''
                sheet['E' + str(ln)] = client.email or ''
                sheet['F' + str(ln)] = client.credit_limit or ''
                sheet['G' + str(ln)] = client.property_payment_term_id.name or ''
                sheet['H' + str(ln)] = client.credit or 0
                sheet['I' + str(ln)] = client.remaining_credit or 0
                ln=ln+1
                
                    
               
                #feuille_courante['F' + str(ln)].fill = couleurBranche
                #feuille_courante['G' + str(ln)].fill = couleurBranche

                total_branche_limite+=client.credit_limit
                total_branche_dette+=client.credit
                total_branche_credit_disponible+=client.remaining_credit
                

                         
          
            ln=ln+1
            sheet['B' + str(ln)] = 'Total'
            feuille_courante['B' + str(ln)].fill = couleurBranche
            sheet['F' + str(ln)] = total_branche_limite
            feuille_courante['F' + str(ln)].fill = couleurBranche
            sheet['H' + str(ln)] = total_branche_dette
            feuille_courante['H' + str(ln)].fill = couleurBranche
            sheet['I' + str(ln)] = total_branche_credit_disponible
            feuille_courante['I' + str(ln)].fill = couleurBranche
            ln=ln+2

        current_date=time.strftime("%Y_%m_%d")
        
        #nomFichier='RAPPORT_credit_limit'+'_'+str(current_date)+'.xlsx' 
        self.name =u'CREDIT_LIMIT_REPORT-'+u'V-'+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")+".xlsx"    
        #output = StringIO()
        #xfile.save(output)

        current_date=time.strftime("%Y_%m_%d")
        chemin=get_module_resource('nh_credit_limit_report', 'static', 'customer_credit_limit')
        nomFichier=chemin+'_'+str(current_date)+'.xlsx'
        xfile.save(nomFichier)
  
        encodedFile=None
        with open(nomFichier, "rb") as f:
            encodedFile = base64.b64encode(f.read())
    
        data=encodedFile
        self.write({'state':'get',
                                'data':data,
                                'name': self.name,
                                })
        
        return {
                    'type' : 'ir.actions.act_window',
                    'res_model' : 'nh_credit_limit_report.credit_limit_report_wizard',
                    'view_mode' : 'form',
                    'view_type' : 'form',
                    'res_id' : self.id,
                    'views' : [(False, 'form')],
                    'target' : 'new',
                    
        }


    
    def get_credit_limit(self, data):

        res = []
        
       
        #_logger.info("Traitement des rapports E/S en PDF :")
        branch_id = data['branch_id']

        #_logger.info("Paramère d'entrées :")
        #_logger.info("-branche :%s",branch_id)
        

       
        if(branch_id and branch_id>0):
            branches=self.env['res.branch'].search([('id','=',branch_id)])
            #_logger.info("branche lu sur le wizard")
        else:
            branches=self.env['res.branch'].search([('id','>',0)])
            #_logger.info("abscence wizard alors, choix de tous les branches")

        context=None
        
        #chargement de la liste des entrepôts
        
        for branche in branches:
            arbre=[]
            total_branche_limite=0
            total_branche_dette=0
            total_branche_credit_disponible=0
            clients=self.env['res.partner'].search(['&',('branch_id','=',branche.id),('customer','=',True)])
            
            #_logger.info("cas de l'branche %s ",branche.name)
            

            for client in clients:
                
                elt= {
                    'code':client.id,
                    'name' : client.name,
                    'phone': client.phone,
                    'email': client.email,
                    'credit_limit':client.credit_limit,
                    'payment_term':client.property_payment_term_id.name,
                    'credit' : client.credit,
                    'remaining_credit' : client.remaining_credit,
                }
                total_branche_limite+=client.credit_limit
                total_branche_dette+=client.credit
                total_branche_credit_disponible+=client.remaining_credit
                arbre.append(elt)
            

            res.append({
                    'branch':branche.name,
                    'total_limit':total_branche_limite,
                    'total_credit':total_branche_dette,
                    'total_remaining_credit':total_branche_credit_disponible,
                    'clients':arbre,
            })
        arbre=[]
        total_branche_limite=0
        total_branche_dette=0
        total_branche_credit_disponible=0
        clients_sans_branches=self.env['res.partner'].search(['&',('customer','=',True),'|',('branch_id','=',False),'|',('branch_id','=',0),('branch_id','=',None)])
        for client in clients_sans_branches:
                
            elt= {
                'code':client.id,
                'name' : client.name,
                'phone': client.phone,
                'email': client.email,
                'credit_limit':client.credit_limit,
                'payment_term':client.property_payment_term_id.name,
                'credit' : client.credit,
                'remaining_credit' : client.remaining_credit,
            }
           
            total_branche_limite+=client.credit_limit
            total_branche_dette+=client.credit
            total_branche_credit_disponible+=client.remaining_credit
            arbre.append(elt)
        res.append({
                    'branch':'Branche inconnue ( Ucknown Branch)',
                    'total_limit':total_branche_limite,
                    'total_credit':total_branche_dette,
                    'total_remaining_credit':total_branche_credit_disponible,
                    'clients':arbre,
            })
                
                    
            
        return res


    def button_print_credit_limit_report(self):
        # datas = {'ids': context.get('active_ids', [])}
        datas = {}
        
        #if self.date_fin < self.date_debut:
        #    raise UserError(_('La date de début doit être plus petite que la date de fin'))
        
        branche_id_val=-1
        if(self.branch_id):
            branche_id_val=self.branch_id.id
        
        datas = {
            
            'branch_id': branche_id_val,
            #'form': vals['form'],
            #'date_debut': self.date_debut,
            #'date_fin': self.date_fin
            
            }

        elements=self.get_credit_limit(datas)
        


        
        datas.update({
            'branch_id':'',
            'ids': self.ids,
            'elements':elements,
            'model': 'nh_credit_limit_report.report_credit_limit'
                        
        })
        #return self.env['report'].get_action(self,'nh_credit_limit_report.report_credit_limit', data=datas)
        #return self.env.ref('nh_credit_limit_report.report_credit_limit').report_action(self, data=datas, config=False)
        return {
                    'type' : 'ir.actions.report',
                    'name':'nh_credit_limit_report.report_credit_limit',
                    'res_model' : 'report.nh_credit_limit_report.report_credit_limit',
                    'model' : 'report.nh_credit_limit_report.report_credit_limit',
                    'report_type':'qweb-pdf',
                    'report_name':'nh_credit_limit_report.report_credit_limit',
                    'data':datas,
                    
                    
        }



        
    
CreditLimitReport()
# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4: