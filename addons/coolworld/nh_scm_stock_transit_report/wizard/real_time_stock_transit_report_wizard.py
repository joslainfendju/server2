# -*- coding: utf-8 -*-
##############################################################################
#
#    ITS/NH-IT
#    Copyright (C) 2018-2019 NH-IT
#
##############################################################################

from io import StringIO, BytesIO
from datetime import datetime


import base64
import datetime
import time

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
from openerp.exceptions import ValidationError

#from openerp.osv import osv,orm
from odoo import api, fields, models,_
import  logging
from odoo.exceptions import UserError

#Autres bibliotheques
import datetime



class data_wizardy(models.TransientModel):
    _name = "nh_scm_stock_transit_report.rt_wizard"



    @api.model
    def _get_company_id(self):
        return self.env.user.company_id


    company_id = fields.Many2one('res.company', string='Company', required="1", default=_get_company_id)
    branch_src_id = fields.Many2one('res.branch', string='Source Branch')
    branch_dest_id = fields.Many2one('res.branch', string='Destination Branch')
    when = fields.Selection([('current', 'At the moment'), ('date', 'On a specific Date')], default='current')
    specific_date = fields.Date('On the')

    filter_by = fields.Selection([('product', 'Product'), ('category', 'Product Category')], string='Filter By',
                                 default='product')
    category_id = fields.Many2one('product.category', string='Category')
    product_ids = fields.Many2many('product.product', string='Products')
    state = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File name', readonly=True)

    @api.constrains('specific_date', 'when')
    def check_constraint(self):
        if (self.when == 'date' and  not self.specific_date):
            raise UserError(_("You most enter a date when you don't choose the current moment"))


    def get_product(self):
        product_pool = self.env['product.product']
        if not self.filter_by:
            product_ids = product_pool.search([('type', '!=', 'service')])
            return product_ids
        elif self.filter_by == 'product' and self.product_ids:
            return self.product_ids
        elif self.filter_by == 'category' and self.category_id:
            product_ids = product_pool.search(
                [('categ_id', 'child_of', self.category_id.id), ('type', '!=', 'service')])
            return product_ids





    @api.multi
    def get_open_quantity(self, product, location):
        start_date = str(self.start_date) + ' 00:00:00'

        return product.with_context({'location': location.id, 'to_date': start_date}).qty_available

    @api.multi
    def get_transit_move_lines(self, products, branch_src=None, branch_dest=None):
        date_format = "%Y-%m-%d %H:%M:%S"

        specific_date = datetime.datetime.strptime(fields.Datetime.now(), date_format).replace(hour=0, minute=0, second=0,
                                                                                       microsecond=0)
        specific_date = specific_date.strftime('%Y-%m-%d') + ' 23:59:59'
        base_domain = []
        en_transit_domain = ['&', ('picking_id', '!=', False), '&', '!', ('picking_id.state', 'in', ('done', 'cancel',)), '&',
                       ('product_id', 'in', [p.id for p in products]),
                       '&', '|', ('picking_id.receipt_from_request_id', '!=', False),
                       ('picking_id.receipt_from_return_request_id', '!=', False),
                       ('date', '<=', specific_date)
                       ]
        if self.when == 'current':
            base_domain = en_transit_domain

        else:
            specific_date = str(self.specific_date) + ' 23:59:59'
            base_domain = ['&', ('picking_id.date', '<=', specific_date),
                           '&',  ('picking_id','!=',False),
                           '&', ('product_id', 'in', [p.id for p in products]),
                           '&', '|',  ('picking_id.receipt_from_request_id', '!=', False), ('picking_id.receipt_from_return_request_id', '!=', False),
                           '&', ('picking_id.state', 'in', ('done', 'cancel',)),
                                ('picking_id.write_date','>=',specific_date),





                           ]
            en_transit_domain = ['&', ('picking_id', '!=', False), '&', '!',
                                 ('picking_id.state', 'in', ('done', 'cancel',)), '&',
                                 ('product_id', 'in', [p.id for p in products]),
                                 '&', '|', ('picking_id.receipt_from_request_id', '!=', False),
                                 ('picking_id.receipt_from_return_request_id', '!=', False),
                                 ('date', '<=', specific_date)
                                 ]
            base_domain = ['|',]+base_domain+en_transit_domain

        if branch_src:
            base_domain = ['&', '|' ,
                                '&' , ('picking_id.receipt_from_request_id','!=', False), ('picking_id.receipt_from_request_id.branch_dest_id','=', branch_src.id),
                                '&',('picking_id.receipt_from_return_request_id','!=', False), ('picking_id.receipt_from_return_request_id.branch_id','=', branch_src.id)
                          ] + base_domain

        if branch_dest:
            base_domain = ['&', '|',
                                '&', ('picking_id.receipt_from_request_id', '!=', False),('picking_id.receipt_from_request_id.branch_id', '=', branch_dest.id),
                                '&', ('picking_id.receipt_from_return_request_id', '!=', False),('picking_id.receipt_from_return_request_id.branch_dest_id', '=', branch_dest.id)
                           ] + base_domain


        logging.info("domaine transit dans stock real transit report est "+ str(base_domain))


        return self.env['stock.move'].search(base_domain)





    def get_product(self):
        product_pool = self.env['product.product']
        if not self.filter_by:
            product_ids = product_pool.search([('type', '!=', 'service')])
            return product_ids
        elif self.filter_by == 'product' and self.product_ids:
            return self.product_ids
        elif self.filter_by == 'category' and self.category_id:
            product_ids = product_pool.search(
                [('categ_id', 'child_of', self.category_id.id), ('type', '!=', 'service')])
            return product_ids





    @api.multi
    def print_pdf(self):

        date_format = "%Y-%m-%d %H:%M:%S"
        specific_date = fields.Datetime.now()

        if self.specific_date:
            specific_date = self.specific_date

        datas = {

            'branch_src_id':self.branch_src_id.name,
            'branch_dest_id': self.branch_dest_id.name,
            'specific_date' : specific_date

        }

        elements =  self.get_json_lines()

        datas.update({

            'ids': self.ids,
            'elements': elements,
            'model': 'nh_scm_stock_transit_report.report_transit'

        })

        return {
            'type': 'ir.actions.report',
            'name': 'nh_scm_stock_transit_report.report_transit',
            'res_model': 'nh_scm_stock_transit_report.report_transit',
            'model': 'report.nh_scm_stock_transit_report.report_transit',
            'report_type': 'qweb-pdf',
            'report_name': 'nh_scm_stock_transit_report.report_transit',
            'data': datas,

        }

    @api.multi
    def exporter_sous_excel(self):

        if (self.when == 'date' and  not self.specific_date):
            raise UserError(_("You most enter a date when you don't choose the current moment"))




        # initialisation du style
        couleurGrandTitre = PatternFill(start_color='2F75B5',
                                        end_color='2F75B5',
                                        fill_type='solid'
                                        )

        couleurTitre = PatternFill(start_color='FFE2EFDA',
                                   end_color='FFE2EFDA',
                                   fill_type='solid')

        couleurBranche = PatternFill(start_color='FF2F75B5',
                                     end_color='FF2F75B5',
                                     fill_type='solid')



        res = []

        # initialisation des données pour Excel
        ln = 8
        path = get_module_resource('nh_scm_stock_transit_report', 'static/template/real_time_stock_transit_report.xlsx')
        xfile = openpyxl.load_workbook(path)
        sheet = xfile.get_sheet_by_name('bilan')



        sheet['D' + str(3)] = self.branch_src_id.name or ""
        sheet['E' + str(3)] = self.branch_dest_id.name or ""

        if self.specific_date:
            sheet['F' + str(3)] = self.specific_date
        else:
            sheet['F' + str(3)] = fields.Datetime.now()

        sheet['G' + str(3)] = self.env.user.name
        sheet['H' + str(3)] = str(datetime.datetime.now())

        # debut du traitement pour chaque entrepôt de la liste des entrepôts

        feuille_courante = xfile.active

        move_lines = self.get_transit_move_lines(self.get_product(), self.branch_src_id, self.branch_dest_id)

        for line in move_lines:
            date_format = "%Y-%m-%d %H:%M:%S"

            datePicking = datetime.datetime.strptime(line.picking_id.date, date_format).replace(hour=0,minute=0,second=0,microsecond=0)

            specific_date = datetime.datetime.strptime(fields.Datetime.now(),date_format).replace(hour=0,minute=0,second=0,microsecond=0)
            #specific_date = specific_date.strftime('%Y-%m-%d') + ' 23:59:59'
            if self.when != 'current':

                specific_date = datetime.datetime.strptime(self.specific_date + ' 00:00:00', date_format).replace(hour=0, minute=0,
                                                                                                       second=0,
                                                                                                       microsecond=0)

            #specific_date = datetime.datetime.strptime(fields.Datetime.now(), date_format)

            dateEnd = None
            if line.picking_id.state in ('done', 'cancel'):
                dateEnd = datetime.datetime.strptime(line.picking_id.write_date, date_format).replace(hour=0,minute=0,second=0,microsecond=0)
                if dateEnd > specific_date:
                    dateEnd = specific_date
            else:
                dateEnd = specific_date

            sheet['J' + str(ln)] =str ( (dateEnd -datePicking).total_seconds()/86400)


            if line.picking_id and line.picking_id.receipt_from_request_id:
                sheet['A' + str(ln)] = line.picking_id.receipt_from_request_id.name
                sheet['C' + str(ln)] = line.picking_id.receipt_from_request_id.branch_dest_id.name
                sheet['D' + str(ln)] = line.picking_id.receipt_from_request_id.branch_id.name+"/"+line.location_dest_id.display_name
            elif  line.picking_id and line.picking_id.receipt_from_return_request_id:
                sheet['A' + str(ln)] = line.picking_id.receipt_from_return_request_id.name
                sheet['C' + str(ln)] = line.picking_id.receipt_from_return_request_id.branch_id.name
                sheet['D' + str(ln)] = line.picking_id.receipt_from_return_request_id.branch_dest_id.name + "/" + line.location_dest_id.display_name


            sheet['B' + str(ln)] = line.date
            if line and line.picking_id:
                sheet['E' + str(ln)] = line.picking_id.name
            sheet['F' + str(ln)] = line.product_id.product_tmpl_id.display_name
           # sheet['G' + str(ln)] = line.lot_id.name
            sheet['G' + str(ln)] = line.product_uom.name
            sheet['H' + str(ln)] = line.product_uom_qty
            sheet['I' + str(ln)] = line.product_id.uom_po_id._compute_quantity(line.product_uom_qty, line.product_uom,
                                                                                rounding_method='HALF-UP') * line.product_id.standard_price
            ln =ln +1




        current_date = time.strftime("%Y_%m_%d")

        # nomFichier='RAPPORT_credit_limit'+'_'+str(current_date)+'.xlsx'
        name = u'STOCK transit REPORT-' + u'V-' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ".xlsx"
        # output = StringIO()
        # xfile.save(output)

        current_date = time.strftime("%Y_%m_%d")
        chemin = get_module_resource('nh_scm_stock_transit_report', 'static', 'template')
        nomFichier = chemin + '_' + str(current_date) + '.xlsx'
        xfile.save(nomFichier)
        self.name = name
        encodedFile = None
        with open(nomFichier, "rb") as f:
            encodedFile = base64.b64encode(f.read())

        data = encodedFile
        self.write({

                    'state': 'get',
                    'data': data,
                    'name': name
                   })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nh_scm_stock_transit_report.rt_wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',

        }

    @api.multi
    def get_json_lines(self):

        res = []
        move_lines = self.get_transit_move_lines(self.get_product(), self.branch_src_id, self.branch_dest_id)
        date_format = "%Y-%m-%d %H:%M:%S"


        for line in move_lines:
            vals = {}

            datePicking = datetime.datetime.strptime(line.picking_id.date, date_format).replace(hour=0, minute=0,
                                                                                                second=0, microsecond=0)
            specific_date = datetime.datetime.strptime(fields.Datetime.now(), date_format).replace(hour=0, minute=0,
                                                                                                   second=0,
                                                                                                   microsecond=0)
            if self.when != 'current':
                specific_date = datetime.datetime.strptime(self.specific_date + ' 00:00:00', date_format).replace(
                    hour=0, minute=0,
                    second=0,
                    microsecond=0)

            dateEnd = None
            if line.picking_id.state in ('done', 'cancel'):
                dateEnd = datetime.datetime.strptime(line.picking_id.write_date, date_format).replace(hour=0, minute=0,
                                                                                                      second=0,
                                                                                                      microsecond=0)
                if dateEnd > specific_date:
                    dateEnd = specific_date
            else:
                dateEnd = specific_date

            vals.update({'days': str(int((dateEnd - datePicking).total_seconds() / 86400))})

            if line.picking_id and line.picking_id.receipt_from_request_id:
                vals.update({'request' :line.picking_id.receipt_from_request_id.name})
                vals.update({'from' :line.picking_id.receipt_from_request_id.branch_dest_id.name})
                vals.update({'to' : line.picking_id.receipt_from_request_id.branch_id.name+"/"+line.location_dest_id.display_name})


            elif  line.picking_id and line.picking_id.receipt_from_return_request_id:
                vals.update({'request': line.picking_id.receipt_from_return_request_id.name})
                vals.update({'from': line.picking_id.receipt_from_return_request_id.branch_id.name})
                vals.update({'to': line.picking_id.receipt_from_return_request_id.branch_dest_id.name + "/" + line.location_dest_id.display_name})

            vals.update({'date': line.date})
            vals.update({'picking':  line.picking_id.name})
            vals.update({'product':  line.product_id.product_tmpl_id.display_name})
            vals.update({'uom':  line.product_uom.name})
            vals.update({'qty':  line.product_uom_qty})
            value = line.product_id.uom_po_id._compute_quantity(line.product_uom_qty, line.product_uom,
                                                                               rounding_method='HALF-UP') * line.product_id.standard_price
            vals.update({'value':  value})


            res.append(vals)
        return res


class real_time_stock_transit_report_excel(models.TransientModel):
    _name = "nh_scm.real_time_stock_transit_report_excel"

    excel_file = fields.Binary('Excel Report')
    file_name = fields.Char('Excel File')

