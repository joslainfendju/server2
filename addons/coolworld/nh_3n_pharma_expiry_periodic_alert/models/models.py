# -*- coding: utf-8 -*-

#bibliotheques pour EXCEL
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
#fin bibliotheques pour EXCEL

#Bibliotheques pour le modèle
from odoo import models, fields, api
#pour les logs
import logging
_logger = logging.getLogger(__name__)


#pour les fichiers
import base64
#Autres bibliotheques
import datetime
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, make_msgid, formatdate, getaddresses, formataddr

import os
import openpyxl
from os import listdir
import openerp.exceptions
from os.path import isfile, join





class Contact(models.Model):
    _name = 'nh_alert_contact.contact'
    _inherit = 'nh_alert_contact.contact'

    alert_expired = fields.Boolean(string='Alert-product-expired',default=False)

    


class NumeroDeSerie(models.Model):
    _inherit=['stock.quant']


   
    def analyserEtAlerter(self):
        #logging.info("##debut de l'analyse des  futurs perimes ou permises###")
        #logging.info("...Chargement des parametres systemes...")

        branches = self.env['res.branch'].search([('id','>',0)])

        for branch in branches:
            have_to_send = False

            parametre_seuil_peremtion=self.sudo().env['ir.config_parameter'].search([('key', '=', 'SEUIL_CRITIQUE_PEREMTION_EN_JOUR')])

            quants = self.env['stock.quant'].search(['&',('location_id.branch_id', '=', branch.id) , ('location_id.scrap_location', '=',False), ('location_id.usage', '=','internal')])

            date_format = "%Y-%m-%d %H:%M:%S"
            seuil_alerte=float(parametre_seuil_peremtion.value)*24*3600#on convertit la valeur du paramètre en secondes
            seuil_perime=0

            #initialisation des données pour Excel
            ln =6
            path = get_module_resource('nh_3n_pharma_expiry_periodic_alert', 'static/template/expiry_template.xlsx')
            xfile = openpyxl.load_workbook(path)
            sheet = xfile.get_sheet_by_name('rapport')
            feuille_courante=xfile.active

            #Initialisation du style
            rouge = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')

            orange = PatternFill(start_color='FFFF8000',
                       end_color='FFFF8000',
                       fill_type='solid')
            sheet['F2'] = branch.name

            sheet['F3'] = str(datetime.datetime.now())
            #Fin Initialisation du style





            #logging.info("...DEBUT DE L'ANALYSE...")
            maintenant=datetime.datetime.today().replace(hour=0,minute=0,second=0,microsecond=0)

            for quant in quants:
                serie=quant.lot_id
                #logging.info("Analyse d'un numero de serie %s",serie.display_name)
                alerter=None
                if(serie and serie.life_date):
                    diff_alerte=None
                    if(serie.alert_date):
                        dateAlerte=datetime.datetime.strptime(serie.alert_date, date_format).replace(hour=0,minute=0,second=0,microsecond=0)
                        diff_alerte=(dateAlerte-maintenant).total_seconds()


                    datePeremption=datetime.datetime.strptime(serie.life_date, date_format).replace(hour=0,minute=0,second=0,microsecond=0)
                    tmp=datePeremption-maintenant
                    diff = tmp.total_seconds()

                    if( (serie.alert_date and diff_alerte>0) or ((not serie.alert_date) and diff>seuil_alerte ) or serie.product_qty==0):
                        alerter=False
                        logging.info("####le numero de serie %s est loin de se perimer et le nbre de secondes restantes est %s secondes/%s secondes",serie.display_name, diff,seuil_alerte)
                    else:
                        alerter=True
                        have_to_send = True
                        sheet['B' + str(ln)] = serie.product_id.code
                        sheet['C' + str(ln)] = serie.product_id.product_tmpl_id.name
                        sheet['D' + str(ln)] = serie.name
                        sheet['E' + str(ln)] = serie.product_id.standard_price
                        sheet['F' + str(ln)] = quant.location_id.complete_name
                        sheet['G' + str(ln)] = serie.product_qty
                        sheet['H' + str(ln)] = serie.product_id.product_tmpl_id.standard_price*serie.product_qty
                        tps=0
                        if(diff>0):
                            tps=diff/86400 #conversion des sécondes en jours
                        sheet['j' + str(ln)] = serie.life_date
                        sheet['k' + str(ln)] = tps


                        if(diff>seuil_perime):
                            #cas de l'alerte
                            #logging.info("####le numero de serie %s est en cas d'alerte et il reste %s heures",serie.display_name,diff)
                            sheet['I' + str(ln)] = 'critique (critical)'
                            feuille_courante['I' + str(ln)].fill = orange


                        else:

                            #logging.info("####le numero de serie %s est perimer et il reste %s heures",serie.display_name,diff)
                            sheet['I' + str(ln)] = 'périmé (expired)'
                            feuille_courante['I' + str(ln)].fill = rouge
                        ln = ln+1

            if have_to_send:
                #output = StringIO()
                #Sauvegarde des données dans le repertoire approprié
                current_date=time.strftime("%Y_%m_%d")
                chemin=get_module_resource('nh_3n_pharma_expiry_periodic_alert', 'static', 'rapport_perimes')
                nomFichier=chemin+'_'+branch.name+'_'+str(current_date)+'.xlsx'
                xfile.save(nomFichier)
                #datas=base64.b64encode(output.getvalue())

                #logging.info("...FIN DE L'ANALYSE...")

                #préparation et envoie du mail

                subject="rapport de l'analyse des périmé du "+time.strftime("%d/%m/%Y %H:%M:%S")
                body_text=""" 
                            Bonjour A vous.
                            Vous trouverez ci-joint, le rapport des articles en situation critique.
                            Cordialement.
                          """
                destinataires = self.env['nh_alert_contact.contact'].search([ '&',('alert_expired','=',True),('active','=',True)])
                destinataires = destinataires.filtered(lambda x: branch.id in [b.id for b in x.branch_ids])
                recipients=[]
                for d in destinataires:
                    recipients.append(d.email)
                if recipients:
                    sender="3n_pharma"
                    outer = MIMEMultipart()
                    outer['Subject'] = subject
                    outer['To'] = str(COMMASPACE.join(recipients))
                    outer['From'] = sender
                    outer['Date'] = formatdate(localtime=True)
                    email_body = MIMEText(body_text)
                    outer.attach(email_body)
                    files=[nomFichier]
                    for f in files:
                        msg=MIMEBase('application', "octet-stream")
                        msg.set_payload(open(f,"rb").read())
                        encoders.encode_base64(msg)
                        msg.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(f))
                        outer.attach(msg)

                    composed=outer.as_string()



                    mail_serverObj=self.env['ir.mail_server']
                    for objMail in mail_serverObj.search([('active','=',True)]):

                        smtp=False
                        try:
                            smtp = mail_serverObj.connect(objMail.smtp_host, objMail.smtp_port, objMail.smtp_user, objMail.smtp_pass, objMail.smtp_encryption or False, objMail.smtp_debug)
                            smtp.sendmail(sender,recipients, composed)

                            smtp.quit()
                        except Exception as e:
                            logging.info("Echec  :%s",e)


        

