# -*- coding: utf-8 -*-
import base64
from datetime import datetime, timedelta
import time

import openpyxl
from openpyxl.styles import PatternFill, Font

from odoo import models, fields, api, exceptions, _

from odoo.modules import get_module_resource
import logging


class wizard(models.TransientModel):
    _name = 'nh_expiry_report.wizard'

    branch_ids = fields.Many2many('res.branch',string="Branch")
    warehouse_ids = fields.Many2many('stock.warehouse', string="Warehouse")
    location_ids = fields.Many2many('stock.location', string="Locations")
    delay = fields.Integer(string='Delay in days', required=True)

    file = fields.Binary('File', readonly=True)
    state = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')
    name = fields.Char(string='File name', readonly=True)

    @api.onchange("branch_ids")
    def _get_warehouse_domain(self):
        """
        This method is for changing the domain of the location once the warehouse changes
        :return:
        """
        if self.branch_ids:
            warehouse_domain = self.env['stock.warehouse'].search([('branch_id', 'in', self.branch_ids.ids)])
            return {'domain': {'warehouse_ids': [('id', 'in', warehouse_domain.ids)]}, 'value': {'warehouse_ids': None}}

    @api.onchange("warehouse_ids")
    def _get_location_domain(self):
        """
        This method is for changing the domain of the location once the warehouse changes
        :return:
        """
        if self.warehouse_ids:
            temp_ids = []
            for wh in self.warehouse_ids:
                temp_ids.append(wh.view_location_id.id)
            loc_domain = self.env['stock.location'].search([('location_id', 'in', temp_ids)])
            return {'domain': {'location_ids': [('id', 'in', loc_domain.ids)]}, 'value': {'location_ids': None}}

    @api.constrains('delay')
    def _check_positive_delay(self):
        """
        The method is to ensure that the delay is positive integer
        :return:
        """
        for r in self:
            if r.delay<0:
                raise exceptions.ValidationError(_("The delay should be greater than 0"))

    @api.multi
    def get_stock_quants(self):
        """
        The methods is for getting the stock quants of the locations
        :return:
        """
        branches = self.branch_ids
        if not branches:
            branches = self.env['res.branch'].search([('id', '>=', 0)])

        warehouses = self.warehouse_ids
        if not warehouses:
            warehouses = self.env['stock.warehouse'].search([('branch_id', 'in', branches.ids)])

        locations = self.location_ids
        if not locations:
            temp_ids = []
            for wh in warehouses:
                temp_ids.append(wh.view_location_id.id)
            locations = self.env['stock.location'].search([('location_id', 'in', temp_ids)])

        current_date = datetime.now()
        max_expiry_date = '{} {}'.format((current_date + timedelta(days=self.delay)).strftime("%Y-%m-%d"), "23:59:59")
        global_stock_quant = self.env['stock.quant'].search([('lot_id.life_date', '<=', max_expiry_date)])
        #logging.info(str(global_stock_quant))
        result = []

        for br in branches:
            result_br = {
                'branch' : br.name,
                'warehouses' : []
            }
            for wh in warehouses:
                if wh.branch_id.id == br.id:
                    result_wh = {
                        'warehouse' : '{}/{}'.format(wh.name, wh.code),
                        'locations' : []
                    }
                    wh_view_loc_id = wh.view_location_id.id
                    for loc in locations:
                        if loc.location_id.id == wh_view_loc_id:
                            result_loc = {
                                'location': loc.display_name,
                                'location_quants': [],
                                'total': 0
                            }
                            for stock_quant in global_stock_quant:
                                if self.check_loc_child_of_loc(loc, stock_quant.location_id):
                                    temp = {
                                        'name_article': stock_quant.product_tmpl_id.name,
                                        'serial_number': stock_quant.lot_id.name,
                                        'unite_value': stock_quant.product_id.standard_price,
                                        'qty': stock_quant.quantity,
                                        'uom': stock_quant.product_uom_id.name,
                                        'location': stock_quant.location_id.display_name
                                    }
                                    if stock_quant.product_id.code:
                                        temp['code_article'] = stock_quant.product_id.code
                                    else:
                                        temp['code_article'] = ''
                                    temp['total_value'] = temp['unite_value']*temp['qty']
                                    result_loc['total'] += temp['total_value']
                                    temp['life_date'] = stock_quant.lot_id.life_date
                                    diff_days = (current_date - datetime.strptime(temp['life_date'] + ".0", '%Y-%m-%d %H:%M:%S.%f')).days

                                    """
                                    if state is True then the article has already expired
                                    """
                                    if diff_days >= 0:
                                        temp['state'] = True
                                        temp['days_to_expiry'] = 0
                                    else:
                                        temp['state'] = False
                                        temp['days_to_expiry'] = -1*diff_days
                                    result_loc['location_quants'].append(temp)
                            if not result_loc['location_quants']:
                                continue
                            result_wh['locations'].append(result_loc)
                    if not result_wh['locations']:
                        continue
                    result_br['warehouses'].append(result_wh)
            if not result_br['warehouses']:
                continue
            result.append(result_br)
        #logging.info(result)
        return result

    def check_loc_child_of_loc(self, parent, child):
        """
        This method returns true in the location 'parent' is really the father of the location 'child'
        or equal to it and false if not
        :param parent:
        :param child:
        :return:
        """
        if parent.parent_left < child.parent_left and parent.parent_left < child.parent_right and parent.parent_right > child.parent_left and parent.parent_right > child.parent_right:
            return True
        if parent.id == child.id:
            return True
        return False

    @api.multi
    def print_pdf(self):
        datas = {
            'delay': self.delay,
            'made_at': datetime.now().strftime("%A %d. %B %Y %H:%M:%S"),
            'result': self.get_stock_quants()
        }

        return {
            'type': 'ir.actions.report',
            'name': 'nh_expiry_report.report',
            'res_model': 'report.nh_expiry_report.report',
            'model': 'report.nh_expiry_report.report',
            'report_type': 'qweb-pdf',
            'report_name': 'nh_expiry_report.report',
            'data': datas,
        }

    @api.multi
    def generate_excel(self):

        font_br = Font(name='Calibri', size=18, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                       color='FF000000')
        font_wh = Font(name='Calibri', size=16, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                       color='FF000000')
        font_loc = Font(name='Calibri', size=14, bold=True, italic=False, vertAlign=None, underline='none',
                        strike=False, color='FF000000')

        datas = {
            'delay': self.delay,
            'made_at': datetime.now().strftime("%A %d. %B %Y %H:%M:%S"),
            'result': self.get_stock_quants()
        }

        path = get_module_resource('nh_expiry_report', 'static/template/template_expired.xlsx')
        xfile = openpyxl.load_workbook(path)
        sheet = xfile['rapport']
        sheet['G2'] = datas['made_at']
        sheet['G3'] = datas['delay']

        result = datas['result']

        cols = [['B', 'code_article', 'CODE'],
                ['C', 'name_article', 'NAME'],
                ['D', 'serial_number', 'SERIAL NUMBER'],
                ['E', 'unite_value', 'UNIT VALUE'],
                ['F', 'location', 'BLOCK LOCATION'],
                ['G', 'qty', 'QUANTITY'],
                ['H', 'uom', 'UNIT OF MEASURE'],
                ['I', 'total_value', 'TOTAL VALUE'],
                ['J', 'state', 'STATE'],
                ['K', 'life_date', 'LIFE DATE'],
                ['L', 'days_to_expiry', 'DAYS BEFORE EXPIRY']]

        row_br_start = 7
        for br in result:
            sheet['B' + str(row_br_start)].font = font_br
            sheet.row_dimensions[row_br_start].height = 29
            sheet['B' + str(row_br_start)] = 'BRANCH : {}'.format(br['branch'])
            row_wh_start = row_br_start + 1
            for wh in br['warehouses']:
                sheet['B' + str(row_wh_start)].font = font_wh
                sheet.row_dimensions[row_wh_start].height = 27
                sheet['B' + str(row_wh_start)] = 'WAREHOUSE : {}'.format(wh['warehouse'])
                row_loc_start = row_wh_start + 1
                for loc in wh['locations']:
                    sheet['B' + str(row_loc_start)].font = font_loc
                    sheet.row_dimensions[row_loc_start].height = 25
                    sheet['B' + str(row_loc_start)] = 'Location : {}'.format(loc['location'])
                    temp = row_loc_start + 1
                    for j in range(len(cols)):
                        sheet[cols[j][0] + str(temp)].font = font_loc
                        sheet[cols[j][0] + str(temp)].fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA',
                                                                         fill_type="solid")
                        sheet.row_dimensions[temp].height = 30
                        sheet[cols[j][0] + str(temp)] = cols[j][2]
                    temp = temp + 1
                    stock_quant = loc['location_quants']
                    for i in range(len(stock_quant)):
                        temp_i = temp + i
                        for col in cols:
                            if col[1] == 'state':
                                if stock_quant[i][col[1]]:
                                    sheet[col[0] + str(temp_i)].fill = PatternFill(start_color='FFFF0000',
                                                                                   end_color='FFFF0000',
                                                                                   fill_type="solid")
                                    sheet[col[0] + str(temp_i)] = 'périmé (expired)'
                                else:
                                    sheet[col[0] + str(temp_i)].fill = PatternFill(start_color='FFFF5733',
                                                                                   end_color='FFFF5733',
                                                                                   fill_type="solid")
                                    sheet[col[0] + str(temp_i)] = 'critique (critical)'
                                continue
                            sheet[col[0] + str(i + row_loc_start + 2)] = stock_quant[i][col[1]]
                    sheet['G' + str(i + row_loc_start + 3)].fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA',
                                                                     fill_type="solid")
                    sheet['G' + str(i + row_loc_start + 3)] = "TOTAL"
                    sheet['H' + str(i + row_loc_start + 3)].fill = PatternFill(start_color='FFE2EFDA',
                                                                               end_color='FFE2EFDA',
                                                                               fill_type="solid")
                    sheet['H' + str(i + row_loc_start + 3)] = loc['total']
                    row_loc_start = i + row_loc_start + 7
                row_wh_start = row_loc_start + 3
            row_br_start = row_wh_start + 4

        current_date = time.strftime("%Y_%m_%d")
        self.name = u'NEARLY_EXPIRY_REPORT-' + u'V-' + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ".xlsx"
        chemin = get_module_resource('nh_expiry_report', 'static/created_excel_reports', '')
        nomFichier = chemin + '_' + str(current_date) + '.xlsx'
        xfile.save(nomFichier)

        encodedFile = None
        with open(nomFichier, "rb") as f:
            encodedFile = base64.b64encode(f.read())

        data = encodedFile
        self.write(
            {
                'state': 'get',
                'file': data,
                'name': self.name,
            }
        )

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nh_expiry_report.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }