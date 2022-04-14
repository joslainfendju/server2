# -*- coding: utf-8 -*-

#bibliotheques pour EXCEL
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import tempfile
from openerp.modules.module import get_module_resource
#from cStringIO import StringIO
#fin bibliotheques pour EXCEL

#Bibliotheques pour le modèle
from odoo import models, fields, api
#pour les logs
import logging
_logger = logging.getLogger(__name__)


#pour les fichiers
import base64
#Autres bibliotheques
import datetime
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, make_msgid, formatdate, getaddresses, formataddr

import os
import openpyxl
from os import listdir
import openerp.exceptions
from os.path import isfile, join





#Fin Bibliotheques pour le modèle

# class nh_3n_pharma_periodic_expirer_alert(models.Model):
#     _name = 'nh_3n_pharma_periodic_expirer_alert.nh_3n_pharma_periodic_expirer_alert'

#     name = fields.Char()
#     value = fields.Integer()
#     value2 = fields.Float(compute="_value_pc", store=True)
#     description = fields.Text()
#
#     @api.depends('value')
#     def _value_pc(self):
#         self.value2 = float(self.value) / 100


class Contact(models.Model):
    _name = 'nh_alert_contact.contact'
    _inherit = 'nh_alert_contact.contact'

    alert_payment = fields.Boolean(string='Alert-payment-overdue',default=False)

    


class AccountInvoice(models.Model):
    _inherit=['account.invoice']


   
    def analyserEtAlerter(self):
        #logging.info("##debut de l'analyse des  futurs perimes ou permises###")
        #logging.info("...Chargement des parametres systemes...")


        parametre_seuil_expiration = self.env['ir.config_parameter'].search(
            [('key', '=', 'SEUIL_CRITIQUE_EXPIRATION_PAYEMENT_EN_JOUR')])

        # logging.info("...Fin Chargement des parametres systemes...")

        # logging.info("...Initialisation des variables de decision...")

        date_format = "%Y-%m-%d"
        seuil_alerte = float(
            parametre_seuil_expiration.value) * 24 * 3600  # on convertit la valeur du paramètre en secondes
        seuil_perime = 0

        # logging.info("...Fin  Initialisation des variables de decision...")

        # initialisation des données pour Excel
        ln = 6
        path = get_module_resource('nh_3n_pharma_payement_alert', 'static/template/overdue_template.xlsx')
        xfile = openpyxl.load_workbook(path)
        sheet = xfile.get_sheet_by_name('rapport')
        feuille_courante = xfile.active

        # Initialisation du style
        rouge = PatternFill(start_color='FFFF0000',
                            end_color='FFFF0000',
                            fill_type='solid')

        orange = PatternFill(start_color='FFFF8000',
                             end_color='FFFF8000',
                             fill_type='solid')


        # Fin Initialisation du style

        # logging.info("...DEBUT DE L'ANALYSE...")
        maintenant = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

        branches = self.env['res.branch'].search([('id', '>', 0)])

        for branch in branches:
            sheet['E3'] = branch.company_id.name
            sheet['E4'] = branch.name
            sheet['E5'] = str(datetime.datetime.now())
            parametre_seuil_peremtion = self.sudo().env['ir.config_parameter'].search(
                [('key', '=', 'SEUIL_CRITIQUE_PEREMTION_EN_JOUR')])

            invoices = self.env['account.invoice'].search(
                ['&',('branch_id','=',branch.id),'&',('state', 'not in', ['draft', 'cancel', 'paid']),('type', 'in', ('out_invoice', 'out_refund'))])
            ln = 8
            for invoice in invoices:
                #logging.info("Analyse d'un numero de invoice %s",invoice.display_name)
                alerter=None
                if(invoice.date_due):


                    dateExpiration = datetime.datetime.strptime(invoice.date_due, date_format).replace(hour=0,minute=0,second=0,microsecond=0)
                    tmp = dateExpiration-maintenant
                    diff = tmp.total_seconds()

                    if(diff>seuil_alerte):
                        alerter=False
                        logging.info("####le numero de invoice %s est loin de se perimer et le nbre de secondes restantes est %s secondes/%s secondes",invoice.number, diff,seuil_alerte)
                    else:
                        alerter=True
                        sheet['B' + str(ln)] = invoice.number
                        sheet['C' + str(ln)] = invoice.partner_id.name
                        sheet['D' + str(ln)] = str(invoice.amount_total)+" "+invoice.currency_id.name
                        sheet['E' + str(ln)] = str(invoice.residual)+" "+invoice.currency_id.name

                        tps=0
                        if(diff>0):
                            tps=diff/86400 #conversion des sécondes en jours
                        sheet['G' + str(ln)] = invoice.date_due
                        sheet['H' + str(ln)] = tps


                        if(diff>seuil_perime):
                            #cas de l'alerte
                            #logging.info("####le numero de invoice %s est en cas d'alerte et il reste %s heures",invoice.display_name,diff)
                            sheet['F' + str(ln)] = 'critique'
                            feuille_courante['F' + str(ln)].fill = orange


                        else:

                            #logging.info("####le numero de invoice %s est perimer et il reste %s heures",invoice.display_name,diff)
                            sheet['F' + str(ln)] = 'échue'
                            feuille_courante['H' + str(ln)].fill = rouge
                        ln=ln+1

            # output = StringIO()
            # Sauvegarde des données dans le repertoire approprié
            current_date = time.strftime("%Y_%m_%d")
            chemin = get_module_resource('nh_3n_pharma_payement_alert', 'static', 'overdue_invoices')
            nomFichier = chemin + '_' + branch.name + '_' + str(current_date) + '.xlsx'
            xfile.save(nomFichier)
            # datas=base64.b64encode(output.getvalue())

            # logging.info("...FIN DE L'ANALYSE...")

            # préparation et envoie du mail

            subject = "rapport de l'analyse des factures échues ou qui vont bientôt l'être "+time.strftime("%d/%m/%Y %H:%M:%S")
            body_text = """ 
                               Bonjour A vous.
                               Vous trouverez ci-joint, le rapport des factures en situation critique.
                               Cordialement.
                             """
            destinataires = self.env['nh_alert_contact.contact'].search(
                ['&', ('alert_expired', '=', True), ('active', '=', True)])
            destinataires = destinataires.filtered(lambda x: branch.id in [b.id for b in x.branch_ids])
            recipients = []
            for d in destinataires:
                recipients.append(d.email)
            if recipients:
                sender = "3n_pharma"
                outer = MIMEMultipart()
                outer['Subject'] = subject
                outer['To'] = str(COMMASPACE.join(recipients))
                outer['From'] = sender
                outer['Date'] = formatdate(localtime=True)
                email_body = MIMEText(body_text)
                outer.attach(email_body)
                files = [nomFichier]
                for f in files:
                    msg = MIMEBase('application', "octet-stream")
                    msg.set_payload(open(f, "rb").read())
                    encoders.encode_base64(msg)
                    msg.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(f))
                    outer.attach(msg)

                composed = outer.as_string()

                mail_serverObj = self.env['ir.mail_server']
                for objMail in mail_serverObj.search([('active', '=', True)]):

                    smtp = False
                    try:
                        smtp = mail_serverObj.connect(objMail.smtp_host, objMail.smtp_port,
                                                      objMail.smtp_user, objMail.smtp_pass,
                                                      objMail.smtp_encryption or False, objMail.smtp_debug)
                        smtp.sendmail(sender, recipients, composed)

                        smtp.quit()
                    except Exception as e:
                        logging.info("Echec  :%s", e)

